# =============================================================================
# Excel Logging Module
# =============================================================================
# Tracks processing runs in Excel workbook with multiple sheets.
# Creates a summary sheet for all devices and individual sheets per device.
#
# Why this exists: Provides a readable log of all time correction runs,
# allowing quick comparison across devices and tracking of processing history.
# =============================================================================
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# ExcelLogger: Log pipeline processing results to Excel workbook
# Creates/updates workbook with summary sheet and per-device history sheets.
# -----------------------------------------------------------------------------
class ExcelLogger:

    def __init__(self, logbook_file: Path):
        self.logbook_file = logbook_file
        self.logbook_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Log a processing run to Excel (updates both summary and device sheets)
    def log_processing_run(self,
                          device_id: str,
                          processing_timestamp: datetime,
                          sd_card_stats: dict,
                          ttn_stats: dict,
                          alignment_stats: dict,
                          output_stats: dict,
                          validation_stats: Optional[dict] = None,
                          issues: str = "None"):
        try:
            # Load or create workbook
            if self.logbook_file.exists():
                wb = openpyxl.load_workbook(self.logbook_file)
            else:
                wb = openpyxl.Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
            
            # Update summary sheet
            self._update_summary_sheet(wb, device_id, processing_timestamp, 
                                      sd_card_stats, alignment_stats, output_stats, issues)
            
            # Update device-specific sheet
            self._update_device_sheet(wb, device_id, processing_timestamp,
                                     sd_card_stats, ttn_stats, alignment_stats, 
                                     output_stats, validation_stats, issues)
            
            # Save workbook
            wb.save(self.logbook_file)
            logger.info(f"Logged processing run to: {self.logbook_file}")
        
        except Exception as e:
            logger.error(f"[ERROR] Failed to log to Excel: {e}")
    
    # Update or create the All_Devices_Summary sheet (one row per device)
    def _update_summary_sheet(self, wb, device_id, timestamp, sd_stats, align_stats, output_stats, issues):
        sheet_name = "All_Devices_Summary"
        
        # Create sheet if doesn't exist
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name, 0)  # Insert at beginning
            
            # Create header
            headers = ['Device', 'Last Processed', 'Status', 'Total Records', 
                      'Drift Mean (days)', 'Drift Max (days)', 'Confidence', 
                      'Output Files', 'Issues']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws = wb[sheet_name]
        
        # Find or create row for this device
        device_row = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, 1).value == device_id:
                device_row = row_idx
                break
        
        if device_row is None:
            device_row = ws.max_row + 1
        
        # Calculate stats
        date_range = sd_stats.get('date_range', (None, None))
        total_records = sd_stats.get('total_lines', 0)
        mean_confidence = align_stats.get('mean_confidence', 0)
        drift_mean = align_stats.get('mean_drift_days', 0) if 'mean_drift_days' in align_stats else 0
        drift_max = align_stats.get('max_drift_days', 0) if 'max_drift_days' in align_stats else 0
        output_files = output_stats.get('files_created', 0)
        
        status = "[OK] Complete" if issues == "None" else "[WARNING] Issues"
        
        # Write data
        ws.cell(device_row, 1, device_id)
        ws.cell(device_row, 2, timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(device_row, 3, status)
        ws.cell(device_row, 4, total_records)
        ws.cell(device_row, 5, round(drift_mean, 2))
        ws.cell(device_row, 6, round(drift_max, 2))
        ws.cell(device_row, 7, round(mean_confidence, 3))
        ws.cell(device_row, 8, output_files)
        ws.cell(device_row, 9, issues)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Update or create device-specific sheet (one row per processing run)
    def _update_device_sheet(self, wb, device_id, timestamp, sd_stats, ttn_stats,
                            align_stats, output_stats, val_stats, issues):
        sheet_name = device_id
        
        # Create sheet if doesn't exist
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            
            # Create header
            headers = ['Run #', 'Date/Time', 'Input Files (OLD)', 'Input Files (NEW)', 
                      'Total Records', 'Date Range', 'Mean Drift', 'Max Drift', 
                      'Confidence', 'Output Files', 'Chronological %', 'Issues/Notes']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws = wb[sheet_name]
        
        # Determine run number
        run_number = ws.max_row  # Header is row 1, so next row is run 2, etc.
        
        # Calculate stats
        total_records = sd_stats.get('total_lines', 0)
        date_range = sd_stats.get('date_range', (None, None))
        date_range_str = f"{date_range[0].strftime('%Y-%m-%d') if date_range[0] else 'N/A'} to {date_range[1].strftime('%Y-%m-%d') if date_range[1] else 'N/A'}"
        
        mean_drift = align_stats.get('mean_drift_days', 0) if 'mean_drift_days' in align_stats else align_stats.get('mean_drift_seconds', 0) / 86400
        max_drift = align_stats.get('max_drift_days', 0) if 'max_drift_days' in align_stats else align_stats.get('max_drift_seconds', 0) / 86400
        mean_confidence = align_stats.get('mean_confidence', 0)
        
        output_files = output_stats.get('files_created', 0)
        
        # TTN file counts
        ttn_old = 0
        ttn_new = 0
        if 'format_breakdown' in ttn_stats:
            ttn_old = ttn_stats['format_breakdown'].get('OLD_TTN', 0)
            ttn_new = ttn_stats['format_breakdown'].get('NEW_SERVER', 0)
        
        # Validation stats
        chronological_pct = 100.0
        if val_stats and 'chronological_files' in val_stats:
            chronological_pct = (val_stats['chronological_files'] / output_files * 100) if output_files > 0 else 100.0
        
        # Write data
        row_data = [
            run_number,
            timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            ttn_old,
            ttn_new,
            total_records,
            date_range_str,
            f"{mean_drift:.2f} days",
            f"{max_drift:.2f} days",
            f"{mean_confidence:.3f}",
            output_files,
            f"{chronological_pct:.1f}%",
            issues
        ]
        
        ws.append(row_data)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Get processing history for a specific device (returns DataFrame or None)
    def get_device_history(self, device_id: str) -> Optional[pd.DataFrame]:
        try:
            if not self.logbook_file.exists():
                return None
            
            # Read device sheet
            df = pd.read_excel(self.logbook_file, sheet_name=device_id)
            return df
        
        except Exception as e:
            logger.warning(f"Could not read device history for {device_id}: {e}")
            return None
    
    # Get summary of all devices (returns DataFrame or None)
    def get_summary(self) -> Optional[pd.DataFrame]:
        try:
            if not self.logbook_file.exists():
                return None
            
            df = pd.read_excel(self.logbook_file, sheet_name="All_Devices_Summary")
            return df
        
        except Exception as e:
            logger.warning(f"Could not read summary: {e}")
            return None
